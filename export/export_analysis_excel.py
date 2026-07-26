"""Export investment_analysis.json to a formatted Excel workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Final

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from project_config import ANALYSIS_JSON_PATH, ANALYSIS_XLSX_PATH


DEFAULT_JSON: Final = ANALYSIS_JSON_PATH
DEFAULT_XLSX: Final = ANALYSIS_XLSX_PATH
HEADER_FILL: Final = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL: Final = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT: Final = Font(color="FFFFFF", bold=True)


class ExcelExportError(Exception):
    """Report an invalid analysis report."""


def _title(value: str) -> str:
    return value.replace("_", " ").title()


def _load_report(json_path: Path) -> dict[str, Any]:
    json_path = json_path.expanduser().resolve()
    if not json_path.is_file():
        raise FileNotFoundError(f"Analysis JSON not found: {json_path}")
    try:
        report = json.loads(json_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise ExcelExportError(f"Could not read valid JSON from {json_path}") from error
    required = {"as_of_date", "risk_profile", "portfolio", "assets", "explanation"}
    if missing := sorted(required.difference(report)):
        raise ExcelExportError(f"Analysis JSON is missing: {', '.join(missing)}")
    return report


def _create_worksheet(workbook: Workbook, title: str) -> Worksheet:
    worksheet = workbook.create_sheet(title)
    if not isinstance(worksheet, Worksheet):
        raise ExcelExportError(f"Could not create worksheet: {title}")
    return worksheet


def _style_header(worksheet: Worksheet, row: int = 1) -> None:
    for cell in worksheet[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _value_width(value: object) -> int:
    if isinstance(value, str):
        return len(value)
    if isinstance(value, bool):
        return len("True") if value else len("False")
    if isinstance(value, (int, float)):
        return len(format(value))
    return 0


def _fit_columns(worksheet: Worksheet, maximum: int = 55) -> None:
    for column_index, cells in enumerate(worksheet.columns, start=1):
        width = max(_value_width(cell.value) for cell in cells)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(width + 2, 11),
            maximum,
        )


def _add_table(worksheet: Worksheet, name: str) -> None:
    if worksheet.max_row < 2:
        return
    end = worksheet.cell(worksheet.max_row, worksheet.max_column).coordinate
    table = Table(displayName=name, ref=f"A1:{end}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _write_records(
    workbook: Workbook,
    sheet_name: str,
    records: list[dict[str, Any]],
    table_name: str,
    fractional_percentages: set[str] | None = None,
    whole_percentages: set[str] | None = None,
    color_scale_columns: set[str] | None = None,
) -> None:
    worksheet = _create_worksheet(workbook, sheet_name)
    if not records:
        worksheet.append(["Status"])
        worksheet.append(["No data available"])
        _style_header(worksheet)
        _fit_columns(worksheet)
        return

    keys = list(records[0])
    worksheet.append([_title(key) for key in keys])
    for record in records:
        worksheet.append([record.get(key) for key in keys])

    fractional_percentages = fractional_percentages or set()
    whole_percentages = whole_percentages or set()
    color_scale_columns = color_scale_columns or set()
    for column_index, key in enumerate(keys, start=1):
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row, column_index)
            if key in fractional_percentages:
                cell.number_format = "0.00%"
            elif key in whole_percentages and isinstance(cell.value, (int, float)):
                cell.value /= 100
                cell.number_format = "0.00%"
        if key in color_scale_columns:
            letter = get_column_letter(column_index)
            worksheet.conditional_formatting.add(
                f"{letter}2:{letter}{worksheet.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B",
                ),
            )

    worksheet.freeze_panes = "A2"
    _style_header(worksheet)
    _add_table(worksheet, table_name)
    _fit_columns(worksheet)


def _write_summary(workbook: Workbook, report: dict[str, Any]) -> None:
    worksheet = workbook.active
    if not isinstance(worksheet, Worksheet):
        raise ExcelExportError("Workbook has no active worksheet")
    worksheet.title = "Summary"
    worksheet.append(["Investment Analysis Report", None])
    worksheet.merge_cells("A1:B1")
    worksheet["A1"].fill = HEADER_FILL
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    worksheet["A1"].alignment = Alignment(horizontal="center")
    worksheet.append(["As Of Date", report["as_of_date"]])
    worksheet.append(["Risk Profile", str(report["risk_profile"]).title()])
    portfolio = report["portfolio"]
    worksheet.append(["Portfolio Name", portfolio["portfolio_name"]])
    worksheet.append(["Assets in the Portfolio", len(report["assets"])])
    worksheet.append(["Expected Return", portfolio["expected_return"]])
    worksheet.append(["Expected Volatility", portfolio["expected_volatility"]])
    worksheet.append(["Concentration", portfolio["concentration"]])
    worksheet.append(["Score", portfolio["score"]])
    worksheet.append(["Portfolio Coverage", portfolio["coverage"]])
    worksheet.append(["Asset Coverage", portfolio["asset_coverage"]])
    worksheet.append([])
    worksheet.append(["Selection", "Highest-scoring portfolio under the existing criteria"])
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row, 1).font = Font(bold=True)
        if worksheet.cell(row, 1).value in {
            "Expected Return", "Expected Volatility", "Concentration",
            "Score", "Portfolio Coverage", "Asset Coverage",
        }:
            worksheet.cell(row, 2).number_format = "0.00%"
    worksheet.freeze_panes = "A2"
    _fit_columns(worksheet)


def _write_warnings(workbook: Workbook, warnings: list[str]) -> None:
    worksheet = _create_worksheet(workbook, "Warnings")
    worksheet.append(["#", "Warning / Limitation"])
    for index, warning in enumerate(warnings, start=1):
        worksheet.append([index, warning])
    _style_header(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 7
    worksheet.column_dimensions["B"].width = 110
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")


def _write_explanation(
    workbook: Workbook,
    explanation: dict[str, Any] | None,
) -> None:
    worksheet = _create_worksheet(workbook, "AI Explanation")
    worksheet.append(["Section", "Content"])
    if not explanation:
        worksheet.append([
            "Status",
            "No AI explanation was generated. Run main.py with --analyze --explain.",
        ])
    else:
        for key, value in explanation.items():
            content = (
                "\n".join(f"• {item}" for item in value)
                if isinstance(value, list)
                else str(value)
            )
            worksheet.append([_title(key), content])
    _style_header(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 110
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row, 1).font = Font(bold=True)
        worksheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")


def export_analysis_to_excel(json_path: Path, xlsx_path: Path) -> Path:
    """Create a workbook containing only the best portfolio and its assets."""
    report = _load_report(json_path)
    workbook = Workbook()
    _write_summary(workbook, report)
    _write_records(
        workbook,
        "Assets",
        report["assets"],
        "BestPortfolioAssets",
        fractional_percentages={"Predicted Return", "Risk Score"},
        color_scale_columns={"Predicted Return", "Risk Score"},
    )
    _write_explanation(workbook, report["explanation"])

    xlsx_path = xlsx_path.expanduser().resolve()
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(xlsx_path)
    return xlsx_path


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export investment_analysis.json to a formatted Excel workbook."
    )
    parser.add_argument("json_path", nargs="?", type=Path, default=DEFAULT_JSON)
    parser.add_argument("xlsx_path", nargs="?", type=Path, default=DEFAULT_XLSX)
    return parser.parse_args()


def main() -> int:
    arguments = parse_arguments()
    try:
        output = export_analysis_to_excel(arguments.json_path, arguments.xlsx_path)
    except (ExcelExportError, OSError, ValueError) as error:
        print(f"Error: {error}")
        return 1
    print(f"Created Excel report: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
