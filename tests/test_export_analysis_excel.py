"""Tests for the JSON-to-Excel report exporter."""

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from export_analysis_excel import export_analysis_to_excel


class ExcelExportTests(unittest.TestCase):
    def test_export_creates_readable_workbook_sheets(self) -> None:
        report = {
            "as_of_date": "2026/07/06",
            "risk_profile": "balanced",
            "diagnostics": {
                "training_samples": 100,
                "validation_samples": 30,
                "validation_mae": 0.05,
                "latest_date": "2026/07/06",
                "target": "forward return",
            },
            "investments": [{
                "rank": 1, "product": "Test Fund", "isin": "TEST",
                "asset_class": "Equity", "currency": "EUR",
                "predicted_return": 0.10, "risk_score": 0.05,
                "model_score": 0.08,
            }],
            "portfolios": [{
                "rank": 1, "portfolio_name": "Test Portfolio",
                "expected_return": 0.08, "expected_volatility": 0.04,
                "concentration": 0.20, "score": 0.05, "coverage": 1.0,
            }],
            "allocations": [{
                "product": "Test Fund", "isin": "TEST",
                "asset_class": "Equity", "currency": "EUR",
                "allocation_percent": 100.0, "predicted_return": 0.10,
                "risk_score": 0.05,
            }],
            "warnings": ["Experimental research only."],
            "explanation": None,
        }
        with tempfile.TemporaryDirectory() as directory:
            json_path = Path(directory) / "report.json"
            xlsx_path = Path(directory) / "report.xlsx"
            json_path.write_text(json.dumps(report), encoding="utf-8")

            export_analysis_to_excel(json_path, xlsx_path)

            workbook = load_workbook(xlsx_path)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Summary", "Investments", "Portfolios", "Allocations",
                    "Warnings", "AI Explanation",
                ],
            )
            self.assertEqual(workbook["Investments"]["B2"].value, "Test Fund")
            self.assertEqual(
                workbook["Allocations"]["E2"].number_format,
                "0.00%",
            )


if __name__ == "__main__":
    unittest.main()
